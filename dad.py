import tkinter as tk
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class GUI:
    def __init__(self, master):
        self.master = master
        master.title("GUI Example")
        
        # Set window size to fullscreen
        width = master.winfo_screenwidth()
        height = master.winfo_screenheight()
        master.geometry(f"{width}x{height}")

        # Text input box
        self.text_input = tk.Entry(master)
        self.text_input.grid(row=0, column=0, padx=5, pady=5)

        # Description input box
        self.description_input = tk.Entry(master)
        self.description_input.grid(row=0, column=1, padx=5, pady=5)

        # Enter button
        self.enter_button = tk.Button(master, text="Enter", command=self.enter_text)
        self.enter_button.grid(row=0, column=2, padx=5, pady=5)

        # History button
        self.history_button = tk.Button(master, text="History", command=self.show_history)
        self.history_button.grid(row=1, column=0, columnspan=2, padx=5, pady=5)

        # Average button
        self.average_button = tk.Button(master, text="Average", command=self.calculate_average)
        self.average_button.grid(row=0, column=3, padx=5, pady=5)

        # Exit button
        self.exit_button = tk.Button(master, text="Exit", command=self.save_and_exit)
        self.exit_button.grid(row=2, column=3, sticky="e", padx=5, pady=5)

        # History text display
        self.history_text = tk.Text(master, height=10, width=50)
        self.history_text.grid(row=2, column=0, columnspan=4, padx=5, pady=5)

        # Load or create Excel workbook
        self.workbook = self.load_or_create_workbook()

        # Load history from workbook
        self.load_history()

    def enter_text(self):
        # Fetch text from input box
        word = self.text_input.get()
        description = self.description_input.get()

        # Get current time
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Record and print current time, word, and description
        entry = f"{current_time}: {word} - {description}"
        self.history_text.insert(tk.END, entry + "\n")

        # Write to Excel workbook
        self.write_to_excel(entry)

        # Clear input fields
        self.text_input.delete(0, tk.END)
        self.description_input.delete(0, tk.END)

    def show_history(self):
        # Display the history
        history = self.history_text.get("1.0", tk.END)
        print(history)

    def calculate_average(self):
        # Calculate average of entered values
        history = self.history_text.get("1.0", tk.END).split("\n")
        values = []
        for entry in history:
            if entry:
                parts = entry.split(":")
                if len(parts) == 2:
                    values.append(float(parts[1].split("-")[0].strip()))
        if values:
            average = sum(values) / len(values)
            print("Average:", average)
        else:
            print("No valid values in history.")

    def save_and_exit(self):
        # Save Excel workbook
        self.workbook.save("history.xlsx")
        
        # Exit the program
        self.master.quit()

    def load_or_create_workbook(self):
        try:
            workbook = openpyxl.load_workbook("history.xlsx")
        except FileNotFoundError:
            workbook = Workbook()
            workbook.active.append(["Date", "Word", "Description"])
        return workbook

    def load_history(self):
        sheet = self.workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] != "Date":  # Skip header row
                entry = f"{row[0]}: {row[1]} - {row[2]}"
                self.history_text.insert(tk.END, entry + "\n")

    def write_to_excel(self, entry):
        sheet = self.workbook.active
        row = [datetime.now()] + entry.split(":")[1].split("-")
        sheet.append(row)

root = tk.Tk()
gui = GUI(root)
root.mainloop()
